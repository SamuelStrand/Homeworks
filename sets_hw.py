import openpyxl

workbook = openpyxl.load_workbook('Temp_hw/data.xlsx')
worksheet = workbook.active

big_arr = []

max_row = worksheet.max_row
max_col = worksheet.max_column

for i in range(1, max_row + 1):
    for j in range(1, max_col+1):
        value = worksheet.cell(row=i, column=j).value
        if value:
            big_arr.append(value)

str_arr1 = big_arr[0:11]
str_arr2 = big_arr[11:22]
str_arr3 = big_arr[22:33]

set1 = set(str_arr1)
set2 = set(str_arr2)
set3 = set(str_arr3)

set_intersection = set2.intersection(set1, set3)
print(set_intersection)

set_differ = set2.difference(set1, set3)
print(set_differ)

set_union = set3.union(set1)
print(set_union)




