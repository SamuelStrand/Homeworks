import openpyxl
from openpyxl.utils   import get_column_letter


class Calc:
    def __init__(self, value1: float, value2: float):
        self.value1 = value1
        self.value2 = value2

    def subtraction(self):
        return self.value1 - self.value2

    def sum(self):
        return self.value1 + self.value2

    def multiply(self):
        return self.value1 * self.value2

    def division(self):
        if self.value2 != 0:
            return self.value1 / self.value2
        else:
            return "Error!"

    def exponentiation(self):
        return self.value1 ** self.value2


obj = Calc(11, 2)
print(obj.exponentiation())


class MyClass:
    def __init__(self, name="_1", index=1, value="_A1"):
        self.name = name
        self.index = index
        self.value = value

    def get_values(self):
        return [self.name, self.index, self.value]


records = []
for i in range(1, 1000 + 1):
    records.append(
        MyClass(
            name=f"_{i}",
            index=i,
            value=f"_A{i}",
        )
    )
print(records)

for i in records:
    print(i.get_values())

book = openpyxl.load_workbook("Temp_hw/data.xlsx")
sheet = book.active
global_rows = []
for row in range(2, sheet.max_row + 1):
    local_row = []
    for column in range(1, sheet.max_column + 1):
        local_row.append(sheet.cell(row=row, column=column).value)
    global_rows.append(local_row)


class Record:
    def __init__(self, row: list):
        self.first = str(row[0])
        self.second = str(row[1])
        self.third = str(row[2])
        self.fourth = str(row[3])
        self.fifth = str(row[4])
        self.sixth = str(row[5])
        self.seventh = str(row[6])


book_new = openpyxl.Workbook()
sheet_new = book_new.active
records = []
for row in global_rows:
    records.append(Record(row=row))

print(records)

row_index = 2
for row in records:
    sheet_new[f"A{row_index}"] = str(row.first)
    sheet_new[f"B{row_index}"] = str(row.second)
    sheet_new[f"C{row_index}"] = str(row.third)
    sheet_new[f"D{row_index}"] = str(row.fourth)
    sheet_new[f"E{row_index}"] = str(row.fifth)
    sheet_new[f"F{row_index}"] = str(row.sixth)
    sheet_new[f"G{row_index}"] = str(row.seventh)
    row_index += 1
book_new.save('Temp_hw/new_table.xlsx')
