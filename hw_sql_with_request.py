import openpyxl
import psycopg2
import requests
from openpyxl.utils import get_column_letter

conn = psycopg2.connect("dbname=new1_postgres_db user=postgres password=123456789")
cur = conn.cursor()

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
                  'AppleWebKit/537.36 (KHTML, like Gecko) '
                  'Chrome/102.0.0.0 Safari/537.36'
}
response = requests.get(url='https://jsonplaceholder.typicode.com/posts/', headers=headers)


# Запись данных из API в Postgres
index = 0
arrays = []
for i in range(1, 10):
    index += 1
    request = response.json()[index]
    user_id = request["userId"]
    # print(request)
    ids = request['id']
    title = request['title']
    body = request['body']
    query_string = f"""
    INSERT INTO public.new1_postgres_tb (id, userid, title, body)
    VALUES ({ids}, {user_id}, '{title}', '{body}')
    """
    cur.execute(query_string)
    conn.commit()


# Чтение массива по условию
cur.execute("""
SELECT * FROM public.new1_postgres_tb
WHERE id > 5
""")
conn.commit()
records = cur.fetchall()
for record in records:
    print(record)

cur.close()
conn.close()

# Запись данных с условием в Excel

book = openpyxl.Workbook()
sheet = book.active

titles = ["Id", "UserId", "Title", "Body"]
index_col = 1
for title in titles:
    sheet[f"{get_column_letter(index_col)}1"] = str(title)
    index_col += 1

row_index = 2
for row in records:
    col_index = 1
    for column in row:
        sheet[f"{get_column_letter(col_index)}{row_index}"] = str(column)
        col_index += 1
    row_index += 1
book.save('Temp_hw/new_table.xlsx')







