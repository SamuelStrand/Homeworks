import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

workbook_1 = openpyxl.load_workbook("Temp_hw/hw1.xlsx")
workbook_2 = openpyxl.load_workbook("Temp_hw/hw2.xlsx")
workbook_3 = openpyxl.load_workbook("Temp_hw/hw3.xlsx")

worksheet = workbook_1.active
worksheet_2 = workbook_2.active
worksheet_3 = workbook_3.active

max_row = worksheet.max_row
print(max_row)
max_column = worksheet.max_column
print(max_column)

max_row_2 = worksheet_2.max_row
print(max_row_2)
max_column_2 = worksheet_2.max_column
print(max_column_2)

max_row_3 = worksheet_3.max_row
print(max_row_3)
max_column_3 = worksheet_3.max_column
print(max_column_3)

workbook_new = Workbook()

ws1 = workbook_new.create_sheet("Sheet_1", 0)
ws1.title = "Sheet_1"

ws2 = workbook_new.create_sheet("Sheet_2", 0)
ws2.title = "Sheet_2"

ws3 = workbook_new.create_sheet("Sheet_3", 0)
ws3.title = "Sheet_3"

var_list_1 = ["A"]
var_list_2 = ["б"]
var_list_3 = ["ц"]

for num in range(1, max_row + 1):
    for name in var_list_1:
        row = num
        # 1 -> A, 3 -> C, 26 -> Z
        col = get_column_letter(var_list_1.index(name) + 1)
        # записываем значение в выбранную ячейку
        ws1[f'{col}{row}'] = str(f"{name}_{num}")
        # print(f'{col}_{row}')

for num in range(1, max_row_2 + 1):
    for name in var_list_2:
        row = num
        # 1 -> A, 3 -> C, 26 -> Z
        col = get_column_letter(var_list_2.index(name) + 2)
        # записываем значение в выбранную ячейку
        ws2[f'{col}{row}'] = str(f"{name}_{num}")
        # print(f'{col}_{row}')

for num in range(1, max_row_3 + 1):
    for name in var_list_3:
        row = num
        # 1 -> A, 3 -> C, 26 -> Z
        col = get_column_letter(var_list_3.index(name) + 3)
        # записываем значение в выбранную ячейку
        ws3[f'{col}{row}'] = str(f"{name}_{num}")
        # print(f'{col}_{row}')

workbook_new.save("Temp_hw/new_hw.xlsx")
